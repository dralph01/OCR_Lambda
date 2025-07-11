import pytesseract
from PIL import Image, ImageFilter
from pdf2image import convert_from_bytes
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import boto3
import base64
import os
import io
import subprocess

# === OCR REGIONS ===
ADDRESS_REGION_1 = (750, 760, 919, 260)    # top envelope
ADDRESS_REGION_2 = (750, 2035, 919, 265)   # bottom envelope

# === S3 CONFIGURATION ===
BUCKET_NAME = 'ocr-envelopes'
S3_KEY_PREFIX = 'ocr-results/'  # Folder in S3 to store Excel files

s3 = boto3.client('s3')

# === TEXT CLEANUP ===
def clean_ocr_text(text):
    lines = text.strip().split('\n')
    cleaned = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if len(line) < 5:
            continue
        if any(c in line for c in '|+*=#[]{}~_<>'):
            continue
        if line.replace(" ", "").isdigit():
            continue
        cleaned.append(line)
    return '\n'.join(cleaned)

# === IMAGE EMBEDDER ===
def insert_image(ws, image_obj, row, col):
    img_path = f"/tmp/temp_{row}_{col}.png"
    image_obj.save(img_path)
    img = XLImage(img_path)
    img.width = 200
    img.height = 80
    ws.row_dimensions[row].height = 75
    ws.add_image(img, ws.cell(row=row, column=col).coordinate)

# === OCR REGION PROCESSOR ===
def process_address_region(ws, base_image, region, filename, label, row):
    x, y, w, h = region
    cropped = base_image.crop((x, y, x + w, y + h))

    # Preprocess: grayscale → threshold → sharpen
    gray = cropped.convert('L')
    bw = gray.point(lambda px: 0 if px < 160 else 255, mode='1')
    sharpened = bw.filter(ImageFilter.SHARPEN)

    # OCR with tuned config
    os.environ['TESSDATA_PREFIX'] = '/opt/'
    config = '--oem 3 --psm 4'
    raw_text = pytesseract.image_to_string(sharpened, config=config)
    text = clean_ocr_text(raw_text)

    if text:
        ws.append([f"{filename} ({label})", text, ""])
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        insert_image(ws, cropped, row, 3)
        return True
    return False

# === FILE CHECKER ===
def s3_file_exists(bucket, key):
    try:
        s3.head_object(Bucket=bucket, Key=key)
        return True
    except:
        return False

# === MAIN HANDLER ===
def lambda_handler(event, context):
    try:
        print("🔎 PATH:", os.environ["PATH"])
        print("🔎 ls /opt/bin:", subprocess.check_output(["ls", "-l", "/opt/bin"]).decode())

        # --- File decoding
        filename = event['headers'].get('filename', 'uploaded_file')
        content_type = event['headers'].get('Content-Type', '').lower()
        ext = os.path.splitext(filename)[-1].lower()

        file_data = base64.b64decode(event['body'])

        # Infer extension if missing or generic
        if not ext or ext == '.uploaded_file':
            if 'pdf' in content_type:
                ext = '.pdf'
            elif 'jpeg' in content_type or 'jpg' in content_type:
                ext = '.jpg'
            elif 'png' in content_type:
                ext = '.png'
            else:
                ext = ''

        # --- Daily Excel filename
        today_str = datetime.utcnow().strftime('%Y-%m-%d')
        excel_key = f"{S3_KEY_PREFIX}ocr_output_{today_str}.xlsx"
        excel_local = f"/tmp/ocr_output_{today_str}.xlsx"

        # --- Load or create Excel
        if s3_file_exists(BUCKET_NAME, excel_key):
            s3.download_file(BUCKET_NAME, excel_key, excel_local)
            wb = load_workbook(excel_local)
            ws = wb.active
            row_index = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Addresses"
            ws.append(['Filename', 'Extracted Address', 'Cropped Preview'])
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 40
            row_index = 2

        # --- Image conversion
        pages = []
        if ext == '.pdf':
            if not file_data.startswith(b'%PDF'):
                return {
                    'statusCode': 400,
                    'body': '❌ Error: The uploaded file is not a valid PDF (missing %PDF header)'
                }
            try:
                print("🔧 Testing direct call to pdftoppm...")
                print(subprocess.check_output(["/opt/bin/pdftoppm", "-v"]).decode())
                pages = convert_from_bytes(file_data, dpi=300, poppler_path="/opt/bin")
            except Exception as e:
                return {
                    'statusCode': 400,
                    'body': f"❌ Error reading PDF file. It may be corrupted or unsupported. Details: {str(e)}"
                }
        elif ext in ('.jpg', '.jpeg', '.png'):
            pages = [Image.open(io.BytesIO(file_data))]
        else:
            return {'statusCode': 400, 'body': '❌ Unsupported file type'}

        # --- Process pages
        for page_num, img in enumerate(pages):
            rotated = img.rotate(-90, expand=True)

            if process_address_region(ws, rotated, ADDRESS_REGION_1, filename, f"page{page_num+1}_top", row_index):
                row_index += 1
            if process_address_region(ws, rotated, ADDRESS_REGION_2, filename, f"page{page_num+1}_bottom", row_index):
                row_index += 1

        # --- Save & upload
        wb.save(excel_local)
        s3.upload_file(excel_local, BUCKET_NAME, excel_key)

        return {
            'statusCode': 200,
            'body': f"✅ Processed and saved to {excel_key}"
        }

    except Exception as e:
        return {'statusCode': 500, 'body': f"❌ Error: {str(e)}"}
